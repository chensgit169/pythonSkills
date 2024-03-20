import openpyxl

from data_base.pyquafu_db.task_database import QuafuTaskDatabase, print_task_info


# 使用 TaskDatabase 类的示例

def from_excel(file_path='./Finish状态任务列表.xlsx',
               db_dir='./.quafu'):
    with QuafuTaskDatabase(db_dir=db_dir) as db:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[wb.sheetnames[0]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            task_id, task_name, priority, status, send_time, finish_time = row
            db.insert_task(task_id=task_id,
                           status=status,
                           task_name=task_name,
                           group_name=None,
                           send_time=send_time,
                           priority=priority,
                           finish_time=finish_time)
        wb.close()


if __name__ == '__main__':
    # from_excel()

    with QuafuTaskDatabase(db_dir='./.quafu') as db:
        tasks = db.find_by_group(None)
        print("Task List:")
        for i, task in enumerate(tasks):
            print(i)
            print_task_info(task)
